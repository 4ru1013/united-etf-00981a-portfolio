import datetime as dt
import pathlib
import re
from typing import Optional, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook

BASE_URL = "https://www.ezmoney.com.tw"
FUND_CODE = "49YTW"  # 00981A 的 fundCode

INFO_URL = f"{BASE_URL}/ETF/Fund/Info?FundCode={FUND_CODE}"
EXPORT_URL = f"{BASE_URL}/ETF/Fund/AssetExcelNPOI?fundCode={FUND_CODE}"


# -----------------------------
# Utilities
# -----------------------------
def ensure_dir(p: pathlib.Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def roc_to_ad_yyyymmdd(roc_date_str: str) -> Optional[str]:
    """
    Accept formats like:
      - 115/01/09
      - 115-01-09
      - 民國115年01月09日
    Return YYYYMMDD in AD, e.g. 20260109
    """
    s = str(roc_date_str).strip()
    if not s:
        return None

    # Normalize
    s = s.replace("年", "/").replace("月", "/").replace("日", "")
    s = s.replace("-", "/")

    m = re.search(r"(\d{2,3})\s*/\s*(\d{1,2})\s*/\s*(\d{1,2})", s)
    if not m:
        return None

    roc_year = int(m.group(1))
    month = int(m.group(2))
    day = int(m.group(3))
    ad_year = roc_year + 1911

    try:
        d = dt.date(ad_year, month, day)
    except ValueError:
        return None
    return d.strftime("%Y%m%d")


def extract_data_date_from_xlsx(xlsx_path: pathlib.Path) -> Optional[str]:
    """
    Try to locate "資料日期" in the first sheet within top-left area,
    then parse ROC date.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    # Scan a reasonable region (top 20 rows, first 10 cols)
    for r in range(1, 21):
        for c in range(1, 11):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            txt = str(v).strip()
            if "資料日期" in txt:
                # Common patterns: "資料日期: 115/01/09"
                # Or the date might be in adjacent cell
                # 1) try parse in same cell
                after = txt.split("資料日期", 1)[-1]
                # remove separators like ":" "："
                after = after.replace("：", ":")
                if ":" in after:
                    after = after.split(":", 1)[-1].strip()

                d = roc_to_ad_yyyymmdd(after)
                if d:
                    wb.close()
                    return d

                # 2) try right cell
                v2 = ws.cell(row=r, column=c + 1).value
                d2 = roc_to_ad_yyyymmdd(v2) if v2 is not None else None
                if d2:
                    wb.close()
                    return d2

    wb.close()
    return None


def find_header_row(df_raw: pd.DataFrame) -> Optional[int]:
    """
    df_raw is read with header=None. We scan first N rows to find a row
    that looks like the real header (contains keywords like 代號/股數/名稱).
    Return row index (0-based) if found.
    """
    must_have_any = ["代號", "股票代號", "標的代號", "證券代號"]
    should_have_any = ["名稱", "標的名稱", "股票名稱", "股名"]
    shares_any = ["股數", "持股股數", "數量", "持有股數"]

    max_scan = min(40, len(df_raw))
    for i in range(max_scan):
        row = df_raw.iloc[i].astype(str).fillna("").tolist()
        row_join = " ".join([x.strip() for x in row if x and x != "nan"]).strip()

        if not row_join or row_join.lower() == "nan":
            continue

        has_code = any(k in row_join for k in must_have_any)
        has_shares = any(k in row_join for k in shares_any)
        has_name = any(k in row_join for k in should_have_any)

        if has_code and has_shares:
            return i
        if has_code and has_name and has_shares:
            return i

    return None


def normalize_colname(s: str) -> str:
    return re.sub(r"\s+", "", str(s)).strip()


def pick_column(cols, candidates):
    """
    Pick first matched column name that contains any candidate substring.
    """
    for cand in candidates:
        for col in cols:
            if cand in normalize_colname(col):
                return col
    return None


def to_int_safe(x) -> int:
    if x is None:
        return 0
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return 0
    s = s.replace(",", "").replace(" ", "")
    try:
        return int(float(s))
    except ValueError:
        return 0


# -----------------------------
# Core: Download, Parse, Diff
# -----------------------------
def download_xlsx(session: requests.Session, out_path: pathlib.Path) -> None:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/130.0 Safari/537.36"
        )
    }

    resp_info = session.get(INFO_URL, headers=headers, timeout=30)
    resp_info.raise_for_status()
    print("[INFO] 打開基金資訊頁成功")

    resp_xlsx = session.get(EXPORT_URL, headers=headers, timeout=60)
    resp_xlsx.raise_for_status()
    print(f"[INFO] 下載 API 回應 Content-Type: {resp_xlsx.headers.get('Content-Type')}")

    out_path.write_bytes(resp_xlsx.content)
    print(f"[OK] Saved XLSX to {out_path}")


def parse_holdings_from_xlsx(xlsx_path: pathlib.Path) -> Tuple[pd.DataFrame, Optional[str]]:
    """
    Return standardized holdings DataFrame:
      columns: code, name, shares
    and data_date (YYYYMMDD) if found.
    """
    data_date = extract_data_date_from_xlsx(xlsx_path)
    if data_date:
        print(f"[INFO] data_date = {data_date}")
    else:
        print("[WARN] 無法從檔案內抓到資料日期，將以今天日期做檔名（但仍可解析持股）")

    df0 = pd.read_excel(xlsx_path, sheet_name=0, header=None, engine="openpyxl")
    header_row = find_header_row(df0)
    if header_row is None:
        preview = df0.head(5).to_string(index=False)
        raise RuntimeError(
            "找不到表頭列（代號/股數）。\n"
            "可能 Excel 格式改了。請檢查 raw 檔案。\n\n"
            f"前 5 列預覽：\n{preview}"
        )

    df = pd.read_excel(
        xlsx_path,
        sheet_name=0,
        header=header_row,
        engine="openpyxl",
    )

    df.columns = [normalize_colname(c) for c in df.columns]

    code_col = pick_column(df.columns, ["代號", "股票代號", "標的代號", "證券代號"])
    name_col = pick_column(df.columns, ["名稱", "標的名稱", "股票名稱", "股名"])
    shares_col = pick_column(df.columns, ["股數", "持股股數", "數量", "持有股數"])

    if not code_col or not shares_col:
        raise RuntimeError(
            f"找不到必要欄位（代號/股數）。\n"
            f"目前欄位：{list(df.columns)}"
        )

    rename_map = {code_col: "code", shares_col: "shares"}
    if name_col:
        rename_map[name_col] = "name"

    df = df[list(rename_map.keys())].copy()
    df = df.rename(columns=rename_map)

    if "name" not in df.columns:
        df["name"] = ""

    # code 永遠當作字串（識別碼，不是數值）
    df["code"] = df["code"].astype("string").str.strip()
    df["name"] = df["name"].astype("string").str.strip()
    df["shares"] = df["shares"].apply(to_int_safe)

    df = df[df["code"].notna()]
    df = df[df["code"].str.len() > 0]
    df = df[~df["code"].str.contains("合計|總計|小計", regex=True, na=False)]

    df = df[df["code"].str.match(r"^[0-9A-Za-z.\-]+$", na=False)]

    df = df.groupby(["code", "name"], as_index=False)["shares"].sum()
    df = df.sort_values("shares", ascending=False).reset_index(drop=True)
    return df, data_date


def compute_diff(prev_df: pd.DataFrame, curr_df: pd.DataFrame) -> pd.DataFrame:
    """
    Return diff dataframe with columns:
      code, name, prev_shares, curr_shares, delta, status
    status: NEW / OUT / UP / DOWN / SAME
    """
    prev = prev_df.copy()
    curr = curr_df.copy()

    # 雙保險：merge key 永遠統一成字串，避免 int64 vs object 造成 merge 失敗
    prev["code"] = prev["code"].astype("string").str.strip()
    curr["code"] = curr["code"].astype("string").str.strip()

    prev = prev.rename(columns={"shares": "prev_shares"})
    curr = curr.rename(columns={"shares": "curr_shares"})

    merged = prev.merge(curr, on=["code"], how="outer", suffixes=("_prev", "_curr"))

    merged["name"] = merged.get("name_curr", "").fillna("")
    if "name_prev" in merged.columns:
        merged.loc[merged["name"].eq("") | merged["name"].isna(), "name"] = merged["name_prev"].fillna("")

    merged["prev_shares"] = merged["prev_shares"].fillna(0).astype(int)
    merged["curr_shares"] = merged["curr_shares"].fillna(0).astype(int)
    merged["delta"] = merged["curr_shares"] - merged["prev_shares"]

    def status_row(r):
        if r["prev_shares"] == 0 and r["curr_shares"] > 0:
            return "NEW"
        if r["prev_shares"] > 0 and r["curr_shares"] == 0:
            return "OUT"
        if r["delta"] > 0:
            return "UP"
        if r["delta"] < 0:
            return "DOWN"
        return "SAME"

    merged["status"] = merged.apply(status_row, axis=1)

    order_map = {"NEW": 0, "UP": 1, "DOWN": 2, "OUT": 3, "SAME": 4}
    merged["order"] = merged["status"].map(order_map).fillna(99)
    merged = merged.sort_values(["order", "delta"], ascending=[True, False]).drop(columns=["order"])

    merged = merged[["code", "name", "prev_shares", "curr_shares", "delta", "status"]].reset_index(drop=True)
    return merged


def write_summary_markdown(diff_df: pd.DataFrame, out_md: pathlib.Path, data_date: str) -> None:
    def top_rows(status, n=15):
        sub = diff_df[diff_df["status"] == status].copy()
        if status in ("DOWN", "OUT"):
            sub = sub.sort_values("delta")
        elif status in ("UP", "NEW"):
            sub = sub.sort_values("delta", ascending=False)
        return sub.head(n)

    lines = []
    lines.append(f"# 00981A Holdings Diff ({data_date})\n")

    counts = diff_df["status"].value_counts().to_dict()
    lines.append("## Summary\n")
    lines.append(
        f"- NEW: {counts.get('NEW',0)} | UP: {counts.get('UP',0)} | DOWN: {counts.get('DOWN',0)} | OUT: {counts.get('OUT',0)} | SAME: {counts.get('SAME',0)}\n"
    )

    for sec, label in [("NEW", "新增持股"), ("UP", "加碼"), ("DOWN", "減碼"), ("OUT", "出清")]:
        sub = top_rows(sec, n=20)
        lines.append(f"## {label} ({sec})\n")
        if sub.empty:
            lines.append("_None_\n")
            continue
        lines.append("| code | name | prev | curr | delta | status |\n")
        lines.append("|---|---|---:|---:|---:|---|\n")
        for _, r in sub.iterrows():
            lines.append(
                f"| {r['code']} | {str(r['name']).replace('|',' ')} | {r['prev_shares']} | {r['curr_shares']} | {r['delta']} | {r['status']} |\n"
            )
        lines.append("\n")

    out_md.write_text("".join(lines), encoding="utf-8")


def main():
    base = pathlib.Path("data")
    raw_dir = base / "raw"
    out_dir = base / "out"
    ensure_dir(raw_dir)
    ensure_dir(out_dir)

    # 1) download raw xlsx to temp name first
    session = requests.Session()
    tmp_path = raw_dir / "00981A_portfolio_tmp.xlsx"
    download_xlsx(session, tmp_path)

    # 2) parse holdings and data_date
    holdings_df, data_date = parse_holdings_from_xlsx(tmp_path)

    # fallback date
    if not data_date:
        data_date = dt.date.today().strftime("%Y%m%d")

    # 3) rename raw file based on data_date
    raw_path = raw_dir / f"00981A_portfolio_{data_date}.xlsx"
    if raw_path.exists():
        tmp_path.unlink(missing_ok=True)
        print(f"[INFO] Raw XLSX already exists: {raw_path}")
    else:
        tmp_path.replace(raw_path)
        print(f"[OK] Raw XLSX moved to: {raw_path}")

    # 4) save standardized holdings
    holdings_path = out_dir / f"00981A_holdings_{data_date}.csv"
    holdings_df.to_csv(holdings_path, index=False, encoding="utf-8-sig")
    print(f"[OK] Saved standardized holdings to {holdings_path}")

    latest_path = out_dir / "00981A_latest.csv"

    # 5) diff vs previous latest
    if latest_path.exists():
        # 關鍵：強制 code 為字串，避免 pandas 自動判型造成 merge 失敗
        prev_df = pd.read_csv(latest_path, dtype={"code": "string"})
        if "code" in prev_df.columns:
            prev_df["code"] = prev_df["code"].str.strip()

        if not {"code", "shares"}.issubset(set(prev_df.columns)):
            print("[WARN] latest.csv 格式不對，將略過 diff。")
        else:
            diff_df = compute_diff(prev_df, holdings_df)
            diff_path = out_dir / f"00981A_diff_{data_date}.csv"
            diff_df.to_csv(diff_path, index=False, encoding="utf-8-sig")
            print(f"[OK] Saved diff to {diff_path}")

            md_path = out_dir / f"00981A_diff_{data_date}.md"
            write_summary_markdown(diff_df, md_path, data_date)
            print(f"[OK] Saved diff summary to {md_path}")
    else:
        print("[INFO] No previous latest.csv found; diff skipped (first run).")

    # 6) update latest
    holdings_df.to_csv(latest_path, index=False, encoding="utf-8-sig")
    print(f"[OK] Updated latest to {latest_path}")


if __name__ == "__main__":
    main()